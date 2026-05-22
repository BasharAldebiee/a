"""
ABS 6401.0 CPI Data Parser
Handles the real ABS Time Series Workbook structure (monthly & quarterly).
Data1 sheet layout:
  Row 1:  Series description headers
  Rows 2-10: Metadata (Unit, Series Type, Data Type, etc.)
  Row 11+: Date + values
  Cols 1-9:   Index Numbers for 9 cities
  Cols 10-18: YoY % Change for 9 cities
  Cols 19-27: MoM (Period) % Change for 9 cities
"""

import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

CITIES = ['Australia', 'Sydney', 'Melbourne', 'Brisbane', 'Adelaide',
          'Perth', 'Hobart', 'Darwin', 'Canberra']

CITY_COL_MAP = {  # 0-indexed column in Data1 sheet
    'index': list(range(1, 10)),   # cols 1-9
    'yoy':   list(range(10, 19)),  # cols 10-18
    'mom':   list(range(19, 28)),  # cols 19-27
}


def load_abs_file(filepath_or_bytes) -> pd.DataFrame:
    """
    Parse ABS 6401.0 Table 1 Excel file.
    Returns clean DataFrame with columns:
      Date, Period, {City}_Index, {City}_YoY, {City}_MoM for each of 9 cities.
    """
    if isinstance(filepath_or_bytes, bytes):
        filepath_or_bytes = BytesIO(filepath_or_bytes)
    wb = load_workbook(filepath_or_bytes, read_only=True, data_only=True)

    if 'Data1' not in wb.sheetnames:
        raise ValueError("Sheet 'Data1' not found. Please upload ABS 6401.0 Table 1 file.")

    ws = wb['Data1']
    rows = list(ws.iter_rows(values_only=True))

    # Data starts at row index 10 (row 11 in Excel)
    records = []
    for r in rows[10:]:
        dt = r[0]
        if dt is None:
            continue
        record = {'Date': pd.Timestamp(dt)}
        for i, city in enumerate(CITIES):
            record[f'{city}_Index'] = _safe_float(r[CITY_COL_MAP['index'][i]])
            record[f'{city}_YoY']   = _safe_float(r[CITY_COL_MAP['yoy'][i]])
            record[f'{city}_MoM']   = _safe_float(r[CITY_COL_MAP['mom'][i]])
        records.append(record)

    df = pd.DataFrame(records).sort_values('Date').reset_index(drop=True)
    df['Period'] = df['Date'].dt.strftime('%b-%Y')
    return df


def _safe_float(val):
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def get_city_df(df: pd.DataFrame, city: str) -> pd.DataFrame:
    """Extract single-city view with renamed columns."""
    cols = ['Date', 'Period', f'{city}_Index', f'{city}_YoY', f'{city}_MoM']
    out = df[cols].copy()
    out.columns = ['Date', 'Period', 'CPI_Index', 'YoY_Pct', 'MoM_Pct']
    return out


def calc_custom_change(df: pd.DataFrame, city: str, start_period: str, end_period: str) -> dict:
    """Calculate CPI movement between any two periods."""
    city_df = get_city_df(df, city)
    s = city_df[city_df['Period'] == start_period]['CPI_Index'].values
    e = city_df[city_df['Period'] == end_period]['CPI_Index'].values
    if len(s) == 0 or len(e) == 0:
        return {}
    start_val, end_val = float(s[0]), float(e[0])
    pct = ((end_val - start_val) / start_val) * 100
    return {
        'start_period': start_period,
        'end_period': end_period,
        'start_val': start_val,
        'end_val': end_val,
        'movement': end_val - start_val,
        'pct_change': pct,
    }
