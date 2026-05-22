"""
Australia Post — CPI Calculator
Data & Pricing Executive Tool  |  Version 2.0
Source: ABS 6401.0 Consumer Price Index, Australia (Monthly Series)
Base period: September 2025 = 100.00
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

from abs_cpi_parser import load_abs_file, get_city_df, calc_custom_change, CITIES

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CPI Calculator | Australia Post",
    page_icon="📮",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── STYLES ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500&family=IBM+Plex+Sans:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }

.ap-header {
    background: #C8102E;
    padding: 24px 32px;
    border-radius: 10px;
    margin-bottom: 24px;
    display: flex; align-items: center; gap: 18px;
    border-left: 6px solid #8B0020;
}
.ap-header h1 { color: white; font-size: 1.75rem; font-weight: 700; margin: 0; letter-spacing: -0.03em; }
.ap-header .sub { color: rgba(255,255,255,0.75); font-size: 0.82rem; margin-top: 5px; font-weight: 300; }
.ap-logo { font-size: 2.5rem; }

.kpi-wrap { display: flex; gap: 12px; margin-bottom: 24px; }
.kpi { background: white; border: 1px solid #e5e5e5; border-radius: 8px;
       padding: 18px 20px; flex: 1; box-shadow: 0 1px 4px rgba(0,0,0,0.05); }
.kpi-label { font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
             letter-spacing: 0.1em; color: #999; margin-bottom: 8px; }
.kpi-val { font-family: 'IBM Plex Mono', monospace; font-size: 1.9rem; font-weight: 500; color: #111; line-height: 1; }
.kpi-sub { font-size: 0.72rem; color: #bbb; margin-top: 5px; }
.kpi-red .kpi-val { color: #C8102E; }
.kpi-green .kpi-val { color: #15803d; }

.signal-box { border-radius: 8px; padding: 14px 20px; margin: 16px 0; display: flex; gap: 14px; align-items: flex-start; }
.signal-icon { font-size: 1.6rem; line-height: 1; }
.signal-title { font-weight: 700; font-size: 0.95rem; color: #111; margin-bottom: 4px; }
.signal-desc { font-size: 0.85rem; color: #444; line-height: 1.5; }
.signal-formula { font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; color: #888; margin-top: 6px; }

.sec-title { font-size: 0.68rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.12em;
             color: #C8102E; border-bottom: 2px solid #C8102E; display: inline-block;
             padding-bottom: 4px; margin-bottom: 14px; }

.badge { display: inline-block; background: #FEE2E2; color: #C8102E; font-size: 0.7rem;
         font-weight: 700; padding: 2px 8px; border-radius: 20px; margin-left: 8px; vertical-align: middle; }
.badge-green { background: #DCFCE7; color: #15803d; }
.badge-yellow { background: #FEF9C3; color: #854D0E; }

.stDownloadButton > button {
    background: #C8102E !important; color: white !important;
    border: none !important; font-weight: 600 !important;
    border-radius: 6px !important; padding: 8px 18px !important; font-size: 0.85rem !important;
}
.stDownloadButton > button:hover { background: #9B0B23 !important; }
section[data-testid="stSidebar"] { background: #FAFAFA; border-right: 1px solid #EEE; }

.abs-note { background: #F0F7FF; border-left: 4px solid #3B82F6; padding: 10px 14px;
            border-radius: 0 6px 6px 0; font-size: 0.8rem; color: #1E3A5F; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)

# ─── DATA LOAD ───────────────────────────────────────────────────────────────

@st.cache_data
def load_bundled_data():
    """Load the real ABS 6401.0 file bundled with the app."""
    try:
        with open("640101.xlsx", "rb") as f:
            return load_abs_file(f.read())
    except FileNotFoundError:
        return None

def try_load_upload(uploaded_file):
    try:
        return load_abs_file(uploaded_file.read())
    except Exception as e:
        st.error(f"Could not parse file: {e}\nPlease ensure this is an ABS 6401.0 Table 1 Time Series workbook.")
        return None

# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

def build_excel_report(city_df, city, start_period, end_period, calc):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CPI Report"

    RED, LIGHT_RED = "C8102E", "FCE8EC"
    GREY, WHITE    = "F5F5F5", "FFFFFF"
    DARK           = "1A1A1A"
    thin = Side(style="thin", color="E0E0E0")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(cell, value, bg=RED, fg="FFFFFF", sz=11, bold=True):
        cell.value = value
        cell.font = Font(name="Calibri", bold=bold, size=sz, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    # Title
    ws.merge_cells("A1:G1")
    hdr_cell(ws["A1"], "Australia Post — CPI Analysis Report", sz=14)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Region: {city}   |   Period: {start_period} → {end_period}   |   ABS 6401.0  (Base: Sep-2025 = 100)"
    ws["A2"].font = Font(name="Calibri", size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # KPI block
    kpis = [
        ("Start Index",  f"{calc['start_val']:.2f}",  "D"),
        ("End Index",    f"{calc['end_val']:.2f}",    "E"),
        ("Movement",     f"{calc['movement']:+.2f} pts", "F"),
        ("% Change",     f"{calc['pct_change']:+.2f}%",  "G"),
    ]
    for label, value, col in kpis:
        ws[f"{col}4"] = label
        ws[f"{col}4"].font = Font(name="Calibri", bold=True, size=8, color="888888")
        ws[f"{col}4"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"] = value
        ws[f"{col}5"].font = Font(name="Calibri", bold=True, size=16,
                                  color=RED if "%" in value or "pts" in value else DARK)
        ws[f"{col}5"].fill = PatternFill("solid", fgColor=LIGHT_RED if "%" in value else GREY)
        ws[f"{col}5"].alignment = Alignment(horizontal="center")
        ws[f"{col}5"].border = bdr
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 28

    # Signal row
    pct = calc['pct_change']
    signal = "🔴 Strong Review Recommended" if abs(pct) >= 4 else ("🟡 Review Warranted" if abs(pct) >= 2 else "🟢 Stable — Monitor Quarterly")
    ws.merge_cells("D6:G6")
    ws["D6"] = f"Pricing Signal: {signal}"
    ws["D6"].font = Font(name="Calibri", bold=True, size=10, color=RED)
    ws["D6"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[6].height = 22

    # Table headers
    headers = ["Period", "CPI Index", "MoM Change (%)", "YoY Change (%)", "vs Start (%)", "Signal"]
    for col_i, h in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_i, value=h)
        hdr_cell(c, h, sz=9)
    ws.row_dimensions[8].height = 20

    # Filter data for selected range
    mask = (city_df['Date'] >= pd.to_datetime(start_period, format="%b-%Y")) & \
           (city_df['Date'] <= pd.to_datetime(end_period,   format="%b-%Y"))
    sub = city_df[mask].copy()
    sub['vs_start'] = ((sub['CPI_Index'] - calc['start_val']) / calc['start_val'] * 100).round(2)

    for row_i, (_, row) in enumerate(sub.iterrows(), 9):
        fill_c = "F9F9F9" if row_i % 2 == 0 else WHITE
        mom_val = row['MoM_Pct']
        yoy_val = row['YoY_Pct']
        vs_val  = row['vs_start']
        row_signal = "▲ Rising" if (mom_val or 0) > 0.5 else ("▼ Easing" if (mom_val or 0) < 0 else "→ Flat")
        vals = [row['Period'], row['CPI_Index'], mom_val, yoy_val, vs_val, row_signal]
        for col_i, v in enumerate(vals, 1):
            c = ws.cell(row=row_i, column=col_i, value=v)
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor=fill_c)
            c.border = bdr
            if col_i in (3, 4, 5) and isinstance(v, (int, float)) and v is not None:
                c.number_format = '+0.00%;-0.00%;0.00%' if col_i != 3 else '0.00%'
        ws.row_dimensions[row_i].height = 15

    for col, w in zip("ABCDEFG", [12, 11, 15, 14, 13, 12, 14]):
        ws.column_dimensions[get_column_letter(ord(col)-64)].width = w

    # Raw data sheet
    ws2 = wb.create_sheet("Raw Data")
    raw_headers = ["Period", "CPI Index", "MoM % Change", "YoY % Change"]
    for i, h in enumerate(raw_headers, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="center")
    for row_i, (_, row) in enumerate(sub.iterrows(), 2):
        ws2.cell(row=row_i, column=1, value=row['Period'])
        ws2.cell(row=row_i, column=2, value=row['CPI_Index'])
        ws2.cell(row=row_i, column=3, value=row['MoM_Pct'])
        ws2.cell(row=row_i, column=4, value=row['YoY_Pct'])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    st.markdown("""
    <div class="ap-header">
        <div class="ap-logo">📮</div>
        <div>
            <h1>CPI Calculator <span style="font-weight:300;font-size:1.1rem;">| Australia Post</span></h1>
            <div class="sub">Data & Pricing Executive Tool &nbsp;·&nbsp; ABS 6401.0 Monthly CPI &nbsp;·&nbsp; Base: Sep-2025 = 100.00 &nbsp;·&nbsp; Latest: <strong>Mar-2026</strong></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── SIDEBAR ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<p class="sec-title">Data Source</p>', unsafe_allow_html=True)

        source = st.radio("", ["Use Bundled ABS File (Mar-2026)", "Upload New ABS File"], label_visibility="collapsed")

        df = None
        if source == "Upload New ABS File":
            up = st.file_uploader("Upload ABS 6401.0 Table 1 (.xlsx)", type=["xlsx"])
            if up:
                df = try_load_upload(up)
                if df is not None:
                    st.success(f"✓ Loaded — {len(df)} months ({df['Period'].iloc[0]} to {df['Period'].iloc[-1]})")
            else:
                st.markdown("""<div class="abs-note">
                Download from <strong>abs.gov.au</strong> → CPI Australia → Data Downloads → <em>640101.xlsx</em>
                </div>""", unsafe_allow_html=True)
        
        if df is None:
            df = load_bundled_data()
            if df is None:
                st.error("Bundled data file not found. Please upload an ABS file.")
                st.stop()

        periods = df['Period'].tolist()

        st.markdown('<br><p class="sec-title">Filters</p>', unsafe_allow_html=True)
        city = st.selectbox("📍 Region", CITIES, index=0)

        col_s, col_e = st.columns(2)
        with col_s:
            start_period = st.selectbox("Start", periods, index=max(0, len(periods)-13),
                                        help="Select the base period")
        with col_e:
            end_period = st.selectbox("End", periods, index=len(periods)-1,
                                      help="Select the comparison period")

        st.markdown('<br><p class="sec-title">Display Options</p>', unsafe_allow_html=True)
        show_all_regions  = st.checkbox("Overlay all regions on chart", value=False)
        show_yoy_chart    = st.checkbox("Show YoY % change chart", value=True)
        show_pricing      = st.checkbox("Pricing guidance panel", value=True)

    # ── VALIDATION ───────────────────────────────────────────────────────────
    s_idx = periods.index(start_period)
    e_idx = periods.index(end_period)
    if s_idx >= e_idx:
        st.error("⚠️ Start period must be before End period.")
        st.stop()

    # ── CALCULATE ────────────────────────────────────────────────────────────
    calc = calc_custom_change(df, city, start_period, end_period)
    city_df = get_city_df(df, city)

    pct  = calc['pct_change']
    mv   = calc['movement']
    sv   = calc['start_val']
    ev   = calc['end_val']

    # Latest ABS YoY (official, from file)
    latest_yoy = city_df.iloc[-1]['YoY_Pct']
    latest_mom = city_df.iloc[-1]['MoM_Pct']
    n_months = e_idx - s_idx

    # ── KPI CARDS ─────────────────────────────────────────────────────────────
    up = pct > 0
    pct_cls   = "kpi-red" if up else "kpi-green"
    arrow     = "▲" if up else "▼"

    st.markdown(f"""
    <div class="kpi-wrap">
      <div class="kpi">
        <div class="kpi-label">CPI Start Value</div>
        <div class="kpi-val">{sv:.2f}</div>
        <div class="kpi-sub">{start_period} · {city}</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">CPI End Value</div>
        <div class="kpi-val">{ev:.2f}</div>
        <div class="kpi-sub">{end_period} · {city}</div>
      </div>
      <div class="kpi {'kpi-red' if mv > 0 else 'kpi-green'}">
        <div class="kpi-label">Index Movement</div>
        <div class="kpi-val">{mv:+.2f}</div>
        <div class="kpi-sub">index points ({n_months} months)</div>
      </div>
      <div class="kpi {pct_cls}">
        <div class="kpi-label">% Change</div>
        <div class="kpi-val">{arrow} {abs(pct):.2f}%</div>
        <div class="kpi-sub">{start_period} → {end_period}</div>
      </div>
      <div class="kpi {'kpi-red' if (latest_yoy or 0)>0 else 'kpi-green'}">
        <div class="kpi-label">Latest YoY (ABS Official)</div>
        <div class="kpi-val">{'▲' if (latest_yoy or 0)>0 else '▼'} {abs(latest_yoy or 0):.1f}%</div>
        <div class="kpi-sub">{df['Period'].iloc[-1]} vs same mth prior yr</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Latest MoM Change</div>
        <div class="kpi-val">{'▲' if (latest_mom or 0)>0 else '▼'} {abs(latest_mom or 0):.1f}%</div>
        <div class="kpi-sub">{df['Period'].iloc[-1]} vs {df['Period'].iloc[-2]}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── PRICING SIGNAL ────────────────────────────────────────────────────────
    if show_pricing:
        st.markdown('<p class="sec-title">Pricing Signal</p>', unsafe_allow_html=True)
        if abs(pct) >= 4.0:
            icon, bg, title = "🔴", "#FEE2E2", "Strong Review Recommended"
            desc = f"CPI has risen <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), exceeding the standard 4% threshold. This data supports a formal pricing adjustment proposal. Estimated impact: for every $1M in revenue, cost base has increased by ~${pct*10000:,.0f}."
            badge = f'<span class="badge">≥ 4% Threshold</span>'
        elif abs(pct) >= 2.0:
            icon, bg, title = "🟡", "#FEF9C3", "Selective Review Warranted"
            desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}). This is within the 2–4% review range — consider selective rate adjustments for CPI-sensitive cost categories (fuel, labour, transport)."
            badge = f'<span class="badge badge-yellow">2–4% Range</span>'
        else:
            icon, bg, title = "🟢", "#DCFCE7", "Stable — Continue Monitoring"
            desc = f"CPI has moved <strong>{pct:+.2f}%</strong> ({start_period}→{end_period}), within acceptable tolerance. No immediate pricing action required. Schedule next review at next quarter end."
            badge = f'<span class="badge badge-green">< 2% Stable</span>'

        formula_note = f"Formula: (({ev:.2f} − {sv:.2f}) ÷ {sv:.2f}) × 100 = <strong>{pct:+.4f}%</strong> &nbsp;|&nbsp; ABS Official YoY for {city}: <strong>{latest_yoy}%</strong>"

        st.markdown(f"""
        <div class="signal-box" style="background:{bg};">
            <div class="signal-icon">{icon}</div>
            <div>
                <div class="signal-title">{title} {badge}</div>
                <div class="signal-desc">{desc}</div>
                <div class="signal-formula">{formula_note}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── CHARTS ───────────────────────────────────────────────────────────────
    mask = (city_df['Date'] >= pd.to_datetime(start_period, format="%b-%Y")) & \
           (city_df['Date'] <= pd.to_datetime(end_period,   format="%b-%Y"))
    sub = city_df[mask].copy()

    chart_col1, chart_col2 = (st.columns(2) if show_yoy_chart else (st.container(), None))

    # Index trend chart
    with chart_col1 if show_yoy_chart else st:
        st.markdown('<p class="sec-title">CPI Index Trend</p>', unsafe_allow_html=True)
        fig = go.Figure()

        if show_all_regions:
            colors_map = {
                'Australia':'#C8102E','Sydney':'#1E40AF','Melbourne':'#15803D',
                'Brisbane':'#92400E','Adelaide':'#7C3AED','Perth':'#0891B2',
                'Hobart':'#B45309','Darwin':'#BE185D','Canberra':'#374151'
            }
            for c in CITIES:
                c_sub = get_city_df(df, c)[mask]
                fig.add_trace(go.Scatter(
                    x=c_sub['Period'], y=c_sub['CPI_Index'], name=c,
                    line=dict(width=2.5 if c == city else 1.5,
                              color=colors_map[c],
                              dash='solid' if c == city else 'dot'),
                    mode="lines",
                ))
            chart_title = f"All Regions — CPI Index"
        else:
            fig.add_trace(go.Scatter(
                x=sub['Period'], y=sub['CPI_Index'], name=city,
                line=dict(color="#C8102E", width=2.5),
                mode="lines+markers",
                marker=dict(size=7, color="#C8102E"),
                fill="tozeroy", fillcolor="rgba(200,16,46,0.06)",
            ))
            # Start/end annotations
            fig.add_annotation(x=start_period, y=sv,
                text=f"<b>{sv:.2f}</b><br>{start_period}",
                showarrow=True, arrowhead=2, ax=-40, ay=-40,
                font=dict(size=10, color="#555"), bgcolor="white", bordercolor="#ddd", borderwidth=1)
            fig.add_annotation(x=end_period, y=ev,
                text=f"<b>{ev:.2f}</b><br>{end_period}",
                showarrow=True, arrowhead=2, ax=40, ay=-40,
                font=dict(size=10, color="#C8102E"), bgcolor="white", bordercolor="#C8102E", borderwidth=1)
            chart_title = f"{city} — CPI Index (Base: Sep-2025 = 100)"

        fig.add_hline(y=100, line_dash="dash", line_color="#ccc", line_width=1,
                      annotation_text="Base: Sep-2025 = 100", annotation_position="right")
        fig.update_layout(
            title=dict(text=chart_title, font=dict(size=13, color="#111"), x=0),
            xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
            yaxis=dict(title="Index", gridcolor="#f5f5f5"),
            plot_bgcolor="white", paper_bgcolor="white",
            hovermode="x unified", height=340,
            legend=dict(orientation="h", y=1.08, x=0),
            margin=dict(l=40, r=10, t=50, b=60),
            font=dict(family="IBM Plex Sans"),
        )
        st.plotly_chart(fig, use_container_width=True)

    # YoY chart
    if show_yoy_chart and chart_col2:
        with chart_col2:
            st.markdown('<p class="sec-title">YoY % Change (ABS Official)</p>', unsafe_allow_html=True)
            yoy_sub = sub.dropna(subset=['YoY_Pct'])
            if yoy_sub.empty:
                st.info("YoY data not available for this period range. ABS only provides YoY from Apr-2025 onward in this file.")
            else:
                fig2 = go.Figure()
                bar_colors = ["#C8102E" if v >= 4 else ("#F59E0B" if v >= 2 else "#16A34A")
                              for v in yoy_sub['YoY_Pct']]
                fig2.add_trace(go.Bar(
                    x=yoy_sub['Period'], y=yoy_sub['YoY_Pct'],
                    marker_color=bar_colors, name="YoY %",
                    text=yoy_sub['YoY_Pct'].apply(lambda x: f"{x:.1f}%"),
                    textposition="outside", textfont=dict(size=10),
                ))
                fig2.add_hline(y=4.0, line_dash="dash", line_color="#C8102E", line_width=1.5,
                               annotation_text="4% review threshold", annotation_position="right",
                               annotation_font=dict(color="#C8102E", size=10))
                fig2.add_hline(y=2.0, line_dash="dot", line_color="#F59E0B", line_width=1,
                               annotation_text="2% monitor level", annotation_position="right",
                               annotation_font=dict(color="#F59E0B", size=9))
                fig2.update_layout(
                    title=dict(text=f"{city} — Annual CPI Change (%)", font=dict(size=13, color="#111"), x=0),
                    xaxis=dict(tickangle=-45, gridcolor="#f5f5f5"),
                    yaxis=dict(title="%", gridcolor="#f5f5f5"),
                    plot_bgcolor="white", paper_bgcolor="white",
                    hovermode="x unified", height=340,
                    margin=dict(l=40, r=10, t=50, b=60),
                    font=dict(family="IBM Plex Sans"),
                    showlegend=False,
                )
                st.plotly_chart(fig2, use_container_width=True)

    # ── RESULTS TABLE ─────────────────────────────────────────────────────────
    st.markdown('<p class="sec-title">Detailed Results Table</p>', unsafe_allow_html=True)

    sub_display = sub.copy()
    sub_display['vs_Start_%'] = ((sub_display['CPI_Index'] - sv) / sv * 100).round(2)
    display_df = sub_display[['Period', 'CPI_Index', 'MoM_Pct', 'YoY_Pct', 'vs_Start_%']].copy()
    display_df.columns = ['Period', 'CPI Index', 'MoM Change (%)', 'YoY Change (%)', f'vs {start_period} (%)']

    def color_pct(val):
        if pd.isna(val) or not isinstance(val, (int, float)): return ''
        if val > 0: return 'color: #C8102E; font-weight: 600'
        if val < 0: return 'color: #15803d; font-weight: 600'
        return ''

    styled = (display_df.style
        .format({'CPI Index': '{:.2f}', 'MoM Change (%)': lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 'YoY Change (%)': lambda x: f'{x:+.1f}%' if pd.notna(x) else '—',
                 f'vs {start_period} (%)': '{:+.2f}%'})
        .map(color_pct, subset=['MoM Change (%)', 'YoY Change (%)', f'vs {start_period} (%)'])
    )
    st.dataframe(styled, use_container_width=True, hide_index=True)

    # ── REGIONAL SNAPSHOT TABLE ───────────────────────────────────────────────
    with st.expander("📊 All Regions Snapshot — Latest Month"):
        latest_row = df.iloc[-1]
        latest_period = latest_row['Period']
        snap_data = []
        for c in CITIES:
            idx_val = latest_row[f'{c}_Index']
            yoy_val = latest_row[f'{c}_YoY']
            mom_val = latest_row[f'{c}_MoM']
            signal = "🔴" if (yoy_val or 0) >= 4 else ("🟡" if (yoy_val or 0) >= 2 else "🟢")
            snap_data.append({
                'Region': c,
                'CPI Index': f"{idx_val:.2f}" if idx_val else '—',
                'YoY % (Official)': f"{yoy_val:+.1f}%" if yoy_val else '—',
                'MoM %': f"{mom_val:+.1f}%" if mom_val else '—',
                'Pricing Signal': signal,
            })
        st.caption(f"Reference period: **{latest_period}** — Source: ABS 6401.0")
        st.dataframe(pd.DataFrame(snap_data), use_container_width=True, hide_index=True)

    # ── EXPORTS ───────────────────────────────────────────────────────────────
    st.markdown('<p class="sec-title">Export</p>', unsafe_allow_html=True)
    ec1, ec2, ec3 = st.columns([1, 1, 3])

    with ec1:
        csv_out = display_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download CSV", data=csv_out,
                           file_name=f"CPI_{city}_{start_period}_{end_period}.csv",
                           mime="text/csv")
    with ec2:
        xlsx_out = build_excel_report(city_df, city, start_period, end_period, calc)
        st.download_button("⬇ Download Excel Report", data=xlsx_out,
                           file_name=f"CPI_Report_{city}_{start_period}_{end_period}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ── FOOTER ────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("""
    <div style="font-size:0.75rem;color:#bbb;text-align:center;padding:6px 0;">
        Australia Post · CPI Calculator · Data & Pricing Executive Tool ·
        Source: <a href="https://www.abs.gov.au/statistics/economy/price-indexes-and-inflation/consumer-price-index-australia/latest-release" target="_blank" style="color:#C8102E;">ABS 6401.0 Monthly CPI (Mar-2026)</a> ·
        Released 29 Apr 2026 · Base Sep-2025 = 100
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
